"""
Utility functions for SCIAN data processing without Prefect decorators.
Used by tests for faster execution.
"""
import re
import unicodedata
from pathlib import Path
from typing import Dict, List, Any
import duckdb
from openpyxl import load_workbook


def to_snake_case(name: str) -> str:
    """Convert sheet name to snake_case for table naming."""
    name = unicodedata.normalize('NFKD', name)
    name = name.encode('ascii', 'ignore').decode('ascii')
    name = re.sub(r'[^\w\s-]', '', name)
    name = re.sub(r'[-\s]+', '_', name)
    return name.lower().strip('_')


def get_sheet_names(xlsx_path: Path) -> List[str]:
    """Get all sheet names from XLSX file using openpyxl."""
    wb = load_workbook(str(xlsx_path), read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    wb.close()
    return sheet_names


def read_sheet_with_merged_cells(xlsx_path: Path, sheet_name: str, header_row: int = 2) -> List[Dict[str, Any]]:
    """Read XLSX sheet handling merged cells and blank rows."""
    wb = load_workbook(str(xlsx_path), data_only=True)
    ws = wb[sheet_name]
    
    headers = []
    for cell in ws[header_row]:
        if cell.value:
            headers.append(str(cell.value).strip())
        else:
            headers.append(f"Column_{cell.column}")
    
    rows = []
    last_codigo_value = None
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=False), start=header_row + 1):
        row_dict = {}
        is_empty_row = True
        
        for col_idx, cell in enumerate(row):
            if col_idx >= len(headers):
                break
                
            header = headers[col_idx]
            value = cell.value
            
            if value is not None and str(value).strip():
                is_empty_row = False
                row_dict[header] = str(value).strip()
            else:
                if header == "Código" and last_codigo_value:
                    row_dict[header] = last_codigo_value
                else:
                    row_dict[header] = None
        
        if not is_empty_row:
            if "Código" in row_dict and row_dict["Código"]:
                last_codigo_value = row_dict["Código"]
            
            rows.append(row_dict)
    
    wb.close()
    return rows


def import_sheet_to_duckdb(
    conn: duckdb.DuckDBPyConnection,
    xlsx_path: Path,
    sheet_name: str,
    table_name: str
) -> int:
    """Import a single sheet from XLSX into DuckDB table."""
    rows = read_sheet_with_merged_cells(xlsx_path, sheet_name)
    
    if not rows:
        return 0
    
    headers = list(rows[0].keys())
    placeholders = ', '.join(['?' for _ in headers])
    columns_def = ', '.join([f'"{h}" VARCHAR' for h in headers])
    
    conn.execute(f"CREATE OR REPLACE TABLE {table_name} ({columns_def})")
    
    for row in rows:
        values = [row.get(h) for h in headers]
        conn.execute(f"INSERT INTO {table_name} VALUES ({placeholders})", values)
    
    result = conn.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()
    return result[0] if result else 0


def import_scian_to_duckdb(xlsx_path: Path, db_path: str = "cache/scian.duckdb") -> Dict[str, int]:
    """Import all sheets from SCIAN XLSX file into DuckDB tables."""
    conn = duckdb.connect(str(db_path))
    
    sheet_names = get_sheet_names(xlsx_path)
    
    results = {}
    for sheet_name in sheet_names:
        table_name = to_snake_case(sheet_name)
        row_count = import_sheet_to_duckdb(conn, xlsx_path, sheet_name, table_name)
        results[table_name] = row_count
    
    conn.close()
    return results
