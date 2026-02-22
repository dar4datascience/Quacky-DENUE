import re
from pathlib import Path
from typing import Dict, List, Any
import duckdb
from prefect import task
from openpyxl import load_workbook


@task(name="to_snake_case")
def to_snake_case(name: str) -> str:
    """
    Convert sheet name to snake_case for table naming.
    
    Args:
        name: Original sheet name
        
    Returns:
        snake_cased table name
    """
    import unicodedata
    
    name = unicodedata.normalize('NFKD', name)
    name = name.encode('ascii', 'ignore').decode('ascii')
    name = re.sub(r'[^\w\s-]', '', name)
    name = re.sub(r'[-\s]+', '_', name)
    return name.lower().strip('_')


@task(name="get_sheet_names")
def get_sheet_names(xlsx_path: Path) -> List[str]:
    """
    Get all sheet names from XLSX file using openpyxl.
    
    Args:
        xlsx_path: Path to XLSX file
        
    Returns:
        List of sheet names
    """
    wb = load_workbook(str(xlsx_path), read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    wb.close()
    
    return sheet_names


@task(name="read_sheet_with_merged_cells")
def read_sheet_with_merged_cells(xlsx_path: Path, sheet_name: str, header_row: int = 2) -> List[Dict[str, Any]]:
    """
    Read XLSX sheet handling merged cells and blank rows.
    
    Args:
        xlsx_path: Path to XLSX file
        sheet_name: Name of sheet to read
        header_row: Row number for headers (1-indexed)
        
    Returns:
        List of dictionaries representing rows
    """
    wb = load_workbook(str(xlsx_path), data_only=True)
    ws = wb[sheet_name]
    
    headers = []
    for cell in ws[header_row]:
        if cell.value:
            headers.append(str(cell.value).strip())
        else:
            headers.append(f"Column_{cell.column}")
    
    rows = []
    last_codigo_value = None
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=False), start=header_row + 1):
        row_dict = {}
        is_empty_row = True
        
        for col_idx, cell in enumerate(row):
            if col_idx >= len(headers):
                break
                
            header = headers[col_idx]
            value = cell.value
            
            if value is not None and str(value).strip():
                is_empty_row = False
                row_dict[header] = str(value).strip()
            else:
                if header == "Código" and last_codigo_value:
                    row_dict[header] = last_codigo_value
                else:
                    row_dict[header] = None
        
        if not is_empty_row:
            if "Código" in row_dict and row_dict["Código"]:
                last_codigo_value = row_dict["Código"]
            
            rows.append(row_dict)
    
    wb.close()
    return rows


@task(name="import_sheet_to_duckdb")
def import_sheet_to_duckdb(
    conn: duckdb.DuckDBPyConnection,
    xlsx_path: Path,
    sheet_name: str,
    table_name: str
) -> int:
    """
    Import a single sheet from XLSX into DuckDB table, handling merged cells and blank rows.
    
    Args:
        conn: DuckDB connection
        xlsx_path: Path to XLSX file
        sheet_name: Name of the sheet to import
        table_name: Name for the DuckDB table
        
    Returns:
        Number of rows imported
    """
    rows = read_sheet_with_merged_cells(xlsx_path, sheet_name)
    
    if not rows:
        return 0
    
    headers = list(rows[0].keys())
    placeholders = ', '.join(['?' for _ in headers])
    columns_def = ', '.join([f'"{h}" VARCHAR' for h in headers])
    
    conn.execute(f"CREATE OR REPLACE TABLE {table_name} ({columns_def})")
    
    for row in rows:
        values = [row.get(h) for h in headers]
        conn.execute(f"INSERT INTO {table_name} VALUES ({placeholders})", values)
    
    result = conn.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()
    return result[0] if result else 0


@task(name="import_scian_to_duckdb")
def import_scian_to_duckdb(
    xlsx_path: Path,
    db_path: str = ":memory:"
) -> Dict[str, int]:
    """
    Import all sheets from SCIAN XLSX file into DuckDB tables.
    
    Args:
        xlsx_path: Path to XLSX file
        db_path: DuckDB database path (default: in-memory)
        
    Returns:
        Dictionary mapping table names to row counts
    """
    conn = duckdb.connect(str(db_path) if db_path != ":memory:" else db_path)
    
    sheet_names = get_sheet_names(xlsx_path)
    
    results = {}
    for sheet_name in sheet_names:
        table_name = to_snake_case(sheet_name)
        row_count = import_sheet_to_duckdb(conn, xlsx_path, sheet_name, table_name)
        results[table_name] = row_count
    
    conn.close()
    return results


@task(name="validate_scian_schema")
def validate_scian_schema(
    db_path: str,
    table_name: str,
    expected_columns: List[str]
) -> bool:
    """
    Validate that a SCIAN table has the expected schema.
    
    Args:
        db_path: DuckDB database path
        table_name: Name of table to validate
        expected_columns: List of expected column names
        
    Returns:
        True if schema is valid
    """
    conn = duckdb.connect(str(db_path))
    
    columns = conn.execute(f"""
        SELECT column_name 
        FROM information_schema.columns 
        WHERE table_name = '{table_name}'
        ORDER BY ordinal_position
    """).fetchall()
    
    actual_columns = [col[0] for col in columns]
    conn.close()
    
    if len(actual_columns) < len(expected_columns):
        raise ValueError(
            f"Table {table_name} has {len(actual_columns)} columns, "
            f"expected at least {len(expected_columns)}"
        )
    
    return True


@task(name="validate_scian_codes")
def validate_scian_codes(db_path: str, table_name: str) -> bool:
    """
    Validate SCIAN code format (2-6 digits).
    
    Args:
        db_path: DuckDB database path
        table_name: Name of table to validate
        
    Returns:
        True if all codes are valid
    """
    conn = duckdb.connect(str(db_path))
    
    invalid_codes = conn.execute(f"""
        SELECT "Código" 
        FROM {table_name}
        WHERE "Código" IS NOT NULL 
        AND NOT regexp_matches("Código", '^[0-9]{{2,6}}$')
        LIMIT 10
    """).fetchall()
    
    conn.close()
    
    if invalid_codes:
        raise ValueError(f"Found invalid SCIAN codes: {invalid_codes}")
    
    return True
